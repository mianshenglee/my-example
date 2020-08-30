import os

import pymysql
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border, Alignment, PatternFill


# 数据库表信息自动生成
class DbInfoGenerator:
    def __init__(self, host, port, user, password, db_name, charset):
        # 初始化数据库操作

        self.db = pymysql.connect(host=host, port=port, user=user,
                                  password=password, database=db_name, charset=charset)
        self.cursor = self.db.cursor()

    def create_file(self, file_path):
        # 获取文件，若文件不存在则创建，存在则删除后重新创建

        if os.path.exists(file_path):
            os.remove(file_path)
        wb = Workbook()
        wb.save(file_path)

    def gen_db_table_info(self, db_name, file_path):
        # 生成数据表信息

        table_info_rows = self.get_table_info(db_name)
        for table_row in table_info_rows:
            print("\n", "*" * 10, "生成表信息：", table_row[1])
            self.gen_table_column_info(table_row, file_path)

    def gen_db_table_info_skip_col(self, db_name, file_path, col_names_skip):
        # 过滤指定列，导出数据表信息到文档

        table_info_rows = self.get_table_info(db_name)
        for table_row in table_info_rows:
            print("\n", "*" * 10, "生成表信息：", table_row[1])
            self.gen_table_column_info(table_row, file_path, col_names_skip)

    def gen_table_column_info(self, table_info_row, file_path, col_names_skip=None):
        # 导出字段信息表到文档

        database_name = table_info_row[0]
        table_name = table_info_row[1]
        table_comment = table_info_row[2]
        # 从数据库获取表信息
        column_info = self.get_table_column_info(database_name, table_name)
        # 写入excel文件
        self.save_column_info_to_excel(table_name, table_comment, column_info, file_path, col_names_skip)

    def save_column_info_to_excel(self, table_name, table_comment, column_info, file_path, col_names_skip):
        # 写入表信息到excel文件

        workbook = openpyxl.load_workbook(file_path)
        # 根据下标获取（下标从0开始）
        sheet = workbook.worksheets[0]
        row_data = [table_name]
        if table_comment:
            row_data = [table_name + "(" + table_comment + ")"]
        sheet.append(row_data)
        rurrent_max_row = sheet.max_row
        # 空行分隔
        sheet.insert_rows(rurrent_max_row)
        # 列名
        col_name_data = ["字段名", "允许为空", "类型", "字段描述"]
        sheet.append(col_name_data)
        for row in column_info:
            if col_names_skip and row[2].lower() in col_names_skip:
                print("#" * 10, "跳过此字段：", row[2])
                continue
            print(row[2] + "," + row[5] + "," + row[10] + "," + row[11])
            row_data = [row[2], row[5], row[10], row[11]]
            sheet.append(row_data)
        workbook.save(file_path)

    def get_table_info(self, db_name):
        # 获取数据表信息

        sql = '''SELECT table_schema, table_name, table_comment FROM information_schema.`TABLES` WHERE TABLE_SCHEMA = %s order by table_name'''
        params = [db_name]
        # 查询数据
        self.cursor.execute(sql, params)
        return self.cursor.fetchall()

    def get_table_column_info(self, database_name, table_name):
        # 获取数据表列信息

        params = [database_name, table_name]
        sql = '''SELECT
            TABLE_SCHEMA AS '库名',
            TABLE_NAME AS '表名',
            COLUMN_NAME AS '列名',
            ORDINAL_POSITION AS '列的排列顺序',
            COLUMN_DEFAULT AS '默认值',
            IS_NULLABLE AS '是否为空',
            DATA_TYPE AS '数据类型',
            CHARACTER_MAXIMUM_LENGTH AS '字符最大长度',
            NUMERIC_PRECISION AS '数值精度(最大位数)',
            NUMERIC_SCALE AS '小数精度',
            COLUMN_TYPE AS '列类型',
            COLUMN_COMMENT AS '注释'
        FROM
            information_schema.`COLUMNS`
        WHERE
            TABLE_SCHEMA = %s
            AND TABLE_NAME = %s
        ORDER BY
            TABLE_NAME,
            ORDINAL_POSITION'''
        # 查询数据
        self.cursor.execute(sql, params)
        return self.cursor.fetchall()

    def set_file_format(self, file_path):
        # 设置表格式

        if not os.path.exists(file_path):
            print("文件不存在，不处理")
            return
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.worksheets[0]
        # 设置各列宽
        sheet.column_dimensions["A"].width = 16
        sheet.column_dimensions["B"].width = 10
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 40

        # 设置表名格式
        max_row = sheet.max_row
        for i in range(1, max_row + 1):
            col1_value = sheet.cell(i, 1).value
            col2_value = sheet.cell(i, 2).value
            # 首列有数据，第2列无数据，则为表名
            if col1_value and not col2_value:
                # 合并表名
                sheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
                # 加粗字体
                font = Font(name="微软雅黑", size=12, bold=True, italic=False, color="000000")
                # 黑色边框
                side_style = Side(style="thin", color="000000")
                border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
                # 居中对齐
                cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # 填充背景色
                p_fill = PatternFill(fill_type="solid", fgColor="BFBFBF")
                # 表名cell格式
                for j in range(1, 5):
                    sheet.cell(i, j).font = font
                    sheet.cell(i, j).border = border
                    sheet.cell(i, j).alignment = cell_alignment
                    sheet.cell(i, j).fill = p_fill
            # 若首列和第2列都有数据，则是表内容
            if col1_value and col2_value:
                # 黑色边框
                side_style = Side(style="thin", color="000000")
                border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)
                # 表名cell格式
                for j in range(1, 5):
                    sheet.cell(i, j).border = border

        workbook.save(file_path)

    def __del__(self):
        # 关闭数据库连接

        self.db.close()
        self.cursor.close()


if __name__ == '__main__':
    excel_path = "E:/pythontest/tableinfo.xlsx"

    host = "localhost"
    port = 3310
    user = "root"
    password = "111111"
    db_name = "zzdx_metabase"
    charset = 'utf8'

    # host = "192.168.6.23"
    # port = 3306
    # user = "root"
    # password = "123456"
    # db_name = "4.0-test"
    # charset = 'utf8'

    col_names_to_skip = ["id", "create_by", "create_time", "update_by", "update_time"
        , "created_by", "created_time", "updated_by", "updated_time", "revision", "sys_create_time"
        , "sys_create_user", "sys_update_time", "sys_update_user", "record_version"]

    dbInfoGenerator = DbInfoGenerator(host, port, user, password, db_name, charset)
    dbInfoGenerator.create_file(excel_path)
    # dbInfoGenerator.gen_db_table_info(db_name, excel_path)
    dbInfoGenerator.gen_db_table_info_skip_col(db_name, excel_path, col_names_to_skip)
    dbInfoGenerator.set_file_format(excel_path)
